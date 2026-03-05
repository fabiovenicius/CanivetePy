from openpyxl import load_workbook


def carregar_dados_excel(arquivo, sheet_name, inicio=0):
    dados = []
    try:
        wb = load_workbook(filename=arquivo)
        for sheet in wb.sheetnames:
            if sheet == sheet_name:
                for r, d in enumerate(wb[sheet].iter_rows(values_only=True)):
                    if r >= inicio:
                        dados.append(d)
    except Exception as err:
        print(err)
    return dados
